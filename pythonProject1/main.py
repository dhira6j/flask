# Import libraries
import qrcode
import datetime
import openpyxl
from tkinter import *

# Generate QR codes
student_data = {'John': '123', 'Mary': '456', 'Bob': '789'}

for student, id in student_data.items():
    qr = qrcode.QRCode(version=1,
                       error_correction=qrcode.constants.ERROR_CORRECT_L,
                       box_size=10,
                       border=4)
    qr.add_data(id)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(f"{student}.png")

# Create GUI window
window = Tk()
window.title("Attendance App")


# Function to record attendance
def record():
    filename = 'attendance.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    today = datetime.date.today()
    column = sheet.max_column + 1

    for student, id in student_data.items():
        cell = sheet.cell(row=1, column=column)
        cell.value = today

        cell = sheet.cell(row=2, column=column)
        cell.value = student

    wb.save(filename)
    label.config(text="Attendance Recorded")


# Label
label = Label(window, text="")
label.pack()

# Button
button = Button(window, text="Record Attendance", command=record)
button.pack()

window.mainloop()
